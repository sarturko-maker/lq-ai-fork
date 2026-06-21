"""XLSX reader — one unit per worksheet, via openpyxl (already a dep, MIT).

Uses ``read_only=True`` (memory-bounded streaming) and ``data_only=True``
(read cached cell values; never recompute formulas — untrusted input is
extracted as literal text only). Each worksheet becomes one unit:
a deterministic ``# <sheet name>`` header line followed by tab-joined,
newline-separated cell text in row-major order. The unit is what the
chunker slices, so the offset invariant holds by construction.
"""

from __future__ import annotations

from io import BytesIO

from app.pipeline.readers._base import (
    OOXML_XLSX_MIME,
    ParsedDocument,
    ParserError,
    build_parsed_document,
    dist_version,
    guard_ooxml,
    ooxml_subtype,
)


class XlsxReader:
    """Extracts per-worksheet text from an ``.xlsx`` workbook."""

    parser_label = "openpyxl"
    mimes = frozenset({OOXML_XLSX_MIME})

    def sniff(self, data: bytes) -> bool:
        return ooxml_subtype(data) == "xlsx"

    def read(self, data: bytes) -> ParsedDocument:
        guard_ooxml(data)
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - install-time error
            raise ParserError("openpyxl is not installed; cannot read XLSX") from exc

        # The container passed sniff + guard_ooxml; a still-malformed workbook
        # must fail closed as ParserError (-> ingestion parse_failed), not
        # escape as a bare openpyxl error (mirrors parsers._run_pymupdf).
        try:
            workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
            try:
                unit_texts: list[str] = []
                for sheet in workbook.worksheets:
                    lines = [f"# {sheet.title}"]
                    for row in sheet.iter_rows(values_only=True):
                        lines.append(
                            "\t".join("" if value is None else str(value) for value in row)
                        )
                    unit_texts.append("\n".join(lines))
            finally:
                workbook.close()
        except Exception as exc:
            raise ParserError(f"openpyxl failed to read XLSX: {exc}") from exc

        return build_parsed_document(
            unit_texts,
            parser=self.parser_label,
            parser_version=f"openpyxl={dist_version('openpyxl')}",
        )
