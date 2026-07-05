"""Tabular / Multi-Document Review endpoints — M3-C2.

Surface (per [PRD §3.14](docs/PRD.md#314-tabular--multi-document-review-m3)):

* ``POST   /api/v1/tabular/preview-cost`` — synchronous cost preview;
  no execution row is created. The UI calls this before showing the
  confirmation modal (Decision C-5).
* ``POST   /api/v1/tabular/execute`` — kick off a tabular execution.
  Creates a :class:`TabularExecution` row at ``status='pending'``,
  enqueues the ARQ worker job on the shared playbook queue (per
  Decision C-3 — same queue as Easy Playbook). Returns 202 + the row.
* ``GET    /api/v1/tabular/executions`` — list the caller's
  executions (paginated, recent-first, soft-deleted excluded).
* ``GET    /api/v1/tabular/executions/{id}`` — full state + grid.
* ``DELETE /api/v1/tabular/executions/{id}`` — soft delete; sets
  ``deleted_at``.
* ``POST   /api/v1/tabular/executions/{id}/cancel`` — set status to
  ``cancelled``; the worker honours this on the next cell-iteration
  boundary check.
* ``POST   /api/v1/tabular/executions/{id}/cells/override`` — the lawyer
  sets a manual override on one agentic-grid cell (ADR-F055 T6 / ADR-F042
  auto-write-then-correct; human action, never an agent tool).
* ``DELETE /api/v1/tabular/executions/{id}/cells/override`` — clear that
  override (revert to the agent value).

Authorization
-------------

Router-level ``Depends(get_active_user)`` gates every endpoint.
Per-endpoint: rows are scoped to ``user_id == caller.id`` (admins
see everything). 404 collapses missing + unauthorized (no
information leakage), matching the M3-A6 Easy Playbook posture.

Column-spec resolution
----------------------

Either ``skill_name`` (resolved at execution start from the live
:class:`SkillRegistry` by reading the skill's
``lq_ai.columns``) OR ``columns`` (ad-hoc spec) is required. The
resolved column list is snapshotted onto the row at request time
(Decision C-1 snapshotting posture: re-rendering the grid a week
later must be honest about what was actually run).

Cost-preview shape
------------------

Per Decision C-5, the preview returns ``cells_count``,
``estimated_tokens``, ``estimated_cost_usd``, and a
``per_tier_breakdown`` informational map. The frontend gates the
big "Run" button behind a confirmation checkbox for previews
above ``CONFIRMATION_THRESHOLD_USD`` ($1.00 default per the prep
doc — the threshold is enforced UI-side, not server-side).
"""

from __future__ import annotations

import csv
import io
import logging
import uuid
from datetime import UTC, datetime
from typing import Annotated, Literal

from fastapi import APIRouter, Depends, HTTPException, Query, Request, Response, status
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.dependencies import ActiveUser, MutatingUser, tenant_admin_visibility
from app.api.projects import _load_visible_project
from app.audit import audit_action
from app.clients.gateway import GatewayClient, get_gateway_client
from app.db.session import get_db
from app.models.document import Document, DocumentChunk
from app.models.tabular import TabularExecution
from app.schemas.tabular import (
    CELL_OVERRIDE_KEYS,
    CellOverrideRequest,
    Citation,
    ColumnSpec,
    TabularExecutionCreate,
    TabularExecutionResponse,
    TabularExecutionSummary,
    TabularPreviewCostRequest,
    TabularPreviewCostResponse,
    TabularResults,
)
from app.skills.registry import MutableSkillRegistry
from app.tabular.cost import estimate_tabular_execution_cost
from app.workers.queue import enqueue_tabular_execution_job

logger = logging.getLogger(__name__)

router = APIRouter(tags=["tabular"])


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _registry(request: Request) -> MutableSkillRegistry:
    """Return the live :class:`MutableSkillRegistry` from app state.

    Mirrors :func:`app.api.skills._registry`. The lifespan handler
    installs ``app.state.skill_registry``; if the handler is somehow
    exercised before that, surface a clear error rather than
    AttributeError.
    """

    holder: MutableSkillRegistry | None = getattr(request.app.state, "skill_registry", None)
    if holder is None:
        from app.errors import InternalError

        raise InternalError(
            message="Skill registry is not initialised; the API process is "
            "not yet ready to serve tabular requests.",
            details={"hint": "lifespan startup did not run"},
        )
    return holder


def _resolve_columns(
    request: Request,
    *,
    skill_name: str | None,
    ad_hoc: list[ColumnSpec] | None,
) -> tuple[str | None, list[ColumnSpec]]:
    """Resolve the request's column spec to (skill_name, columns).

    Either ``skill_name`` OR ``ad_hoc`` must be supplied. When
    ``skill_name`` is given, look up the registered skill's
    ``lq_ai.columns`` and snapshot the list. When ``ad_hoc`` is
    given, use it directly. Both-or-neither is a 400.

    Returns the resolved skill_name (None for ad-hoc) + a non-empty
    list of :class:`ColumnSpec`. Raises HTTPException(400) on
    validation errors and HTTPException(404) on unknown skill /
    skill-without-columns.
    """

    if skill_name and ad_hoc:
        raise HTTPException(
            status_code=400,
            detail="provide either skill_name or columns, not both",
        )
    if not skill_name and not ad_hoc:
        raise HTTPException(
            status_code=400,
            detail="provide either skill_name or columns",
        )

    if skill_name:
        record = _registry(request).current().get(skill_name)
        if record is None:
            raise HTTPException(status_code=404, detail=f"skill {skill_name!r} not found")
        skill_columns = record.frontmatter.lq_ai.columns
        if not skill_columns:
            raise HTTPException(
                status_code=400,
                detail=f"skill {skill_name!r} has no columns (not a table-mode skill)",
            )
        # Re-validate through the wire-side ColumnSpec so any
        # skill-side fields the wire schema doesn't honor are
        # stripped consistently.
        resolved = [ColumnSpec.model_validate(col.model_dump()) for col in skill_columns]
        # Bake the skill-level ensemble_verification fallback into each
        # column at this single resolution point (Decision C-1
        # snapshotting posture) so both preview-cost and execute see a
        # consistent snapshot. A column that didn't declare its own
        # value (None) inherits the skill-level value; an explicit
        # True/False per column is left untouched.
        skill_ensemble = record.frontmatter.lq_ai.ensemble_verification
        if skill_ensemble is not None:
            for col in resolved:
                if col.ensemble_verification is None:
                    col.ensemble_verification = skill_ensemble
        return skill_name, resolved

    # ad_hoc branch — the request-validation min_length=1 on columns
    # in TabularExecutionCreate already covers the empty case.
    assert ad_hoc is not None  # for the type checker
    return None, ad_hoc


async def _load_caller_owned_documents(
    db: AsyncSession,
    *,
    document_ids: list[uuid.UUID],
    user: ActiveUser,
) -> list[Document]:
    """Load documents the caller has access to via the file ownership chain.

    Mirrors the M3-A6 Easy Playbook helper. The visibility rule:
    each :class:`Document` has a parent :class:`File` whose
    ``owner_id`` must match the caller (org-admins see all; the
    OPERATOR is excluded from that widening — ADR-F064 D2) AND which
    has not been soft-deleted. Missing / cross-owner / soft-deleted
    documents collapse into "not found" at the caller — running
    tabular extraction over documents the user no longer "has"
    would surprise on the audit trail.
    """

    from app.models.file import File as FileModel

    stmt = (
        select(Document)
        .where(Document.id.in_(document_ids))
        .join(FileModel, Document.file_id == FileModel.id)
        .where(FileModel.deleted_at.is_(None))
    )
    if not tenant_admin_visibility(user):
        stmt = stmt.where(FileModel.owner_id == user.id)
    rows = (await db.execute(stmt)).scalars().all()
    return list(rows)


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


@router.post(
    "/tabular/preview-cost",
    response_model=TabularPreviewCostResponse,
    summary="Preview the cost of a proposed tabular execution.",
)
async def preview_tabular_cost(
    body: TabularPreviewCostRequest,
    request: Request,
    user: ActiveUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    gateway: Annotated[GatewayClient, Depends(get_gateway_client)],
) -> TabularPreviewCostResponse:
    """Synchronous cost preview — no execution row is created.

    The UI calls this before showing the confirmation modal so the
    operator sees the cell-count + estimated cost + per-tier
    breakdown (Decision C-5). The estimator's rolling average
    (M2-E2 pattern) caches per ~5 minutes in-process; rapid re-previews
    during column-spec editing don't thrash the DB.

    Ensemble columns preview higher (Donna #6): the estimator adds one
    ensemble-pass cost (N judge calls) per ensemble cell. We fetch the
    gateway's resolved ensemble config so the premium reflects the
    deployment's actual judge models + default; the config is
    process-cached and ``get_citation_engine_ensemble_config`` already
    returns ``None`` on any gateway error / unreachable, so an
    unconfigured or down gateway simply yields no premium (graceful
    degrade — the preview stays a non-fatal pre-flight).

    Authorization is the router-level ``get_active_user`` gate — any
    authenticated caller may preview costs against their own document
    selection. The document-ownership check is enforced at execute
    time; preview doesn't load the documents.
    """

    _ = user  # gate-only; no per-user logic on preview

    _, columns = _resolve_columns(
        request,
        skill_name=body.skill_name,
        ad_hoc=body.columns,
    )

    # ``columns`` is the RESOLVED spec (skill-level ensemble fallback
    # already baked in by ``_resolve_columns``); the estimator only needs
    # to apply the column-value-or-deployment-default resolution on top.
    ensemble_config = await gateway.get_citation_engine_ensemble_config()

    return await estimate_tabular_execution_cost(
        db,
        document_ids=list(body.document_ids),
        columns=columns,
        ensemble_config=ensemble_config,
    )


@router.post(
    "/tabular/execute",
    response_model=TabularExecutionResponse,
    status_code=status.HTTP_202_ACCEPTED,
    summary="Start a tabular execution.",
)
async def create_tabular_execution(
    body: TabularExecutionCreate,
    request: Request,
    user: MutatingUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> TabularExecutionResponse:
    """Kick off a tabular execution against the supplied document corpus.

    Authorization: the caller must own every document in
    ``document_ids``. Cross-user or non-existent ids collapse into a
    single 404 (no information leakage), matching the M3-A6 Easy
    Playbook posture.

    Creates a :class:`TabularExecution` row at ``status='pending'``
    and enqueues the ARQ worker job on the shared playbook queue
    (per Decision C-3). Returns 202 immediately with the row; the
    result-view polls
    ``GET /api/v1/tabular/executions/{id}`` until status reaches
    a terminal state.

    Per the M3-C2 quality bar, "execution completed" does NOT mean
    every cell is correct — the result-view's per-cell citation
    drawer is where the user-attorney validates extractions before
    relying on them.
    """

    documents = await _load_caller_owned_documents(
        db,
        document_ids=list(body.document_ids),
        user=user,
    )
    if len(documents) != len(body.document_ids):
        raise HTTPException(status_code=404, detail="one or more documents not found")

    resolved_skill_name, columns = _resolve_columns(
        request,
        skill_name=body.skill_name,
        ad_hoc=body.columns,
    )

    # Snapshot the column spec to the row's JSONB so re-rendering a
    # week later honours Decision C-1's snapshotting invariant.
    columns_json = [col.model_dump(mode="json") for col in columns]

    execution = TabularExecution(
        user_id=user.id,
        skill_name=resolved_skill_name,
        status="pending",
        document_ids=list(body.document_ids),
        columns=columns_json,
        cost_estimate_usd=body.confirmed_cost_usd,
    )
    db.add(execution)
    await db.flush()

    await audit_action(
        db,
        user_id=user.id,
        action="tabular.execution_started",
        resource_type="tabular_execution",
        resource_id=str(execution.id),
        request=request,
        details={
            "document_count": len(body.document_ids),
            "column_count": len(columns),
            "skill_name": resolved_skill_name,
            "confirmed_cost_usd": (
                str(body.confirmed_cost_usd) if body.confirmed_cost_usd is not None else None
            ),
        },
    )
    await db.commit()
    await db.refresh(execution)

    enqueued = await enqueue_tabular_execution_job(execution.id)
    logger.info(
        "tabular_execution_started",
        extra={
            "event": "tabular_execution_started",
            "user_id": str(user.id),
            "execution_id": str(execution.id),
            "enqueued": enqueued,
            "document_count": len(body.document_ids),
            "column_count": len(columns),
        },
    )

    return await _to_response(db, execution)


@router.get(
    "/tabular/executions",
    response_model=list[TabularExecutionSummary],
    summary="List the caller's tabular executions.",
)
async def list_tabular_executions(
    user: ActiveUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    limit: int = 50,
) -> list[TabularExecutionSummary]:
    """Return the caller's tabular executions, recent-first, soft-deleted excluded.

    Org-admins see all executions; non-admins — and the OPERATOR, which is
    excluded from the admin widening (ADR-F064 D2) — see only their own
    (matches the Easy Playbook list posture).
    """

    stmt = select(TabularExecution).where(TabularExecution.deleted_at.is_(None))
    if not tenant_admin_visibility(user):
        stmt = stmt.where(
            or_(TabularExecution.user_id == user.id, TabularExecution.user_id.is_(None))
        )
    stmt = stmt.order_by(TabularExecution.created_at.desc()).limit(limit)
    rows = (await db.execute(stmt)).scalars().all()
    return [_to_summary(row) for row in rows]


# A matter accrues few grids; cap the listing defensively so the response stays bounded.
_MATTER_GRIDS_LIMIT = 200


@router.get(
    "/tabular/matters/{project_id}/grids",
    response_model=list[TabularExecutionSummary],
    summary="List a matter's agentic grids (F2 Tabular T7).",
)
async def list_matter_grids(
    project_id: uuid.UUID,
    user: ActiveUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> list[TabularExecutionSummary]:
    """The matter's agentic grids for the cockpit Grids tab (ADR-F055), recent-first.

    Owner-scoped through the matter (``_load_visible_project`` 404s on miss /
    cross-user / archived — no existence leak). Only ``mode='agentic'`` rows
    (the frozen linear executor's rows never surface here) and soft-deleted rows
    are excluded.
    """
    project = await _load_visible_project(db, project_id, user.id)
    stmt = (
        select(TabularExecution)
        .where(
            TabularExecution.project_id == project.id,
            TabularExecution.mode == "agentic",
            TabularExecution.deleted_at.is_(None),
        )
        .order_by(TabularExecution.created_at.desc())
        .limit(_MATTER_GRIDS_LIMIT)
    )
    rows = (await db.execute(stmt)).scalars().all()
    return [_to_summary(row) for row in rows]


@router.get(
    "/tabular/executions/{execution_id}",
    response_model=TabularExecutionResponse,
    summary="Get a tabular execution by id.",
)
async def get_tabular_execution(
    execution_id: uuid.UUID,
    user: ActiveUser,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> TabularExecutionResponse:
    """Return one tabular execution — full state + grid results.

    Authorization: caller must be the row's ``user_id`` OR an admin.
    Cross-user / missing / soft-deleted rows collapse into 404.
    """

    row = await _load_caller_execution(db, execution_id=execution_id, user=user)
    response = await _to_response(db, row)
    await _enrich_cell_citations(db, response)
    return response


@router.delete(
    "/tabular/executions/{execution_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Soft-delete a tabular execution.",
    response_class=Response,
)
async def delete_tabular_execution(
    execution_id: uuid.UUID,
    user: MutatingUser,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> Response:
    """Soft-delete (sets ``deleted_at``). Already-deleted rows return 404."""

    row = await _load_caller_execution(db, execution_id=execution_id, user=user)
    row.deleted_at = datetime.now(UTC)
    await audit_action(
        db,
        user_id=user.id,
        action="tabular.execution_deleted",
        resource_type="tabular_execution",
        resource_id=str(execution_id),
        request=request,
    )
    await db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.post(
    "/tabular/executions/{execution_id}/cancel",
    response_model=TabularExecutionResponse,
    summary="Cancel a pending or running tabular execution.",
)
async def cancel_tabular_execution(
    execution_id: uuid.UUID,
    user: MutatingUser,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> TabularExecutionResponse:
    """Set status to ``cancelled``.

    The worker's per-cell loop checks the row status at each cell
    boundary; on cancellation it stops accepting new cells and the
    aggregate node writes whatever partial results landed. Already-
    terminal rows (``completed`` / ``failed`` / ``cancelled``) are
    a 409 — operators wanting a clean re-run start a new execution
    via ``POST /tabular/execute``.
    """

    row = await _load_caller_execution(db, execution_id=execution_id, user=user)
    if row.status in {"completed", "failed", "cancelled"}:
        raise HTTPException(
            status_code=409,
            detail=f"execution already in terminal status {row.status!r}",
        )
    row.status = "cancelled"
    row.completed_at = datetime.now(UTC)
    await audit_action(
        db,
        user_id=user.id,
        action="tabular.execution_cancelled",
        resource_type="tabular_execution",
        resource_id=str(execution_id),
        request=request,
    )
    await db.commit()
    await db.refresh(row)
    return await _to_response(db, row)


# ---------------------------------------------------------------------------
# Cell override — the lawyer's manual correction (ADR-F055 T6 / ADR-F042)
# ---------------------------------------------------------------------------
#
# "System proposes, user owns" for the matter-tier grid: the agent fills cells
# (``record_tabular_row`` / ``update_tabular_cells``, guarded tools), and the
# lawyer corrects them here. This is a HUMAN-write endpoint, never an agent
# tool — mirrors ``matter_memory.create_matter_correction`` (ADR-F042 §B2).
# The override rides the ``results`` JSONB (no migration); the agent write path
# (``tabular_tool._upsert_row``) preserves the override_* keys so a re-pull can
# never clobber it (the structural "human wins" guarantee).


async def _load_owned_agentic_execution_locked(
    db: AsyncSession,
    *,
    execution_id: uuid.UUID,
    user: ActiveUser,
) -> TabularExecution:
    """Load + row-lock one agentic grid the caller OWNS, for a results write.

    Strict owner scope (``user_id == caller.id`` — no admin bypass: an override
    carries the lawyer's provenance). Pinned to ``mode == 'agentic'`` so it can
    never mutate a frozen linear-executor row (ADR-F001). ``FOR UPDATE`` guards
    the results read-modify-write against concurrent fan-out writers.
    404 collapses missing / cross-user / linear / soft-deleted (no leak).
    """

    stmt = (
        select(TabularExecution)
        .where(
            TabularExecution.id == execution_id,
            TabularExecution.user_id == user.id,
            TabularExecution.mode == "agentic",
            TabularExecution.deleted_at.is_(None),
        )
        .with_for_update()
    )
    row = (await db.execute(stmt)).scalar_one_or_none()
    if row is None:
        raise HTTPException(status_code=404, detail="execution not found")
    return row


def _set_cell_override(
    results: dict | None,
    *,
    document_id: str,
    column_name: str,
    override_value: str,
    override_note: str | None,
    overridden_by: str,
    overridden_at: str,
) -> dict | None:
    """Return a NEW results dict with the override merged onto one cell.

    ``None`` if the (document_id, column_name) cell does not exist (→ 404: you
    can only correct a cell the agent attempted; a finalized grid attempted
    every cell). Reassign-don't-mutate → SQLAlchemy flags the JSONB dirty."""

    rows = [dict(r) for r in (results or {}).get("rows", []) if isinstance(r, dict)]
    for row in rows:
        if row.get("document_id") == document_id:
            cells = dict(row.get("cells", {}))
            cell = cells.get(column_name)
            if not isinstance(cell, dict):
                return None
            new_cell = dict(cell)
            new_cell["override_value"] = override_value
            new_cell["override_note"] = override_note
            new_cell["overridden_by"] = overridden_by
            new_cell["overridden_at"] = overridden_at
            cells[column_name] = new_cell
            row["cells"] = cells
            return {"rows": rows}
    return None


def _clear_cell_override(
    results: dict | None,
    *,
    document_id: str,
    column_name: str,
) -> dict | None:
    """Return a NEW results dict with the override_* keys stripped from one cell
    (revert to the agent value). ``None`` if the cell does not exist (→ 404).
    Idempotent — clearing an un-overridden cell is a harmless no-op."""

    rows = [dict(r) for r in (results or {}).get("rows", []) if isinstance(r, dict)]
    for row in rows:
        if row.get("document_id") == document_id:
            cells = dict(row.get("cells", {}))
            cell = cells.get(column_name)
            if not isinstance(cell, dict):
                return None
            cells[column_name] = {k: v for k, v in cell.items() if k not in CELL_OVERRIDE_KEYS}
            row["cells"] = cells
            return {"rows": rows}
    return None


@router.post(
    "/tabular/executions/{execution_id}/cells/override",
    response_model=TabularExecutionResponse,
    summary="Set a lawyer's manual override on one grid cell.",
)
async def override_tabular_cell(
    execution_id: uuid.UUID,
    payload: CellOverrideRequest,
    user: MutatingUser,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> TabularExecutionResponse:
    """Set a human override on one cell (ADR-F055 T6 / ADR-F042).

    The override shadows the agent's value in display; the agent's value +
    citations stay recorded underneath and the agent write path can never
    clobber the override. Owner-scoped agentic grids only (cross-user /
    linear → 404). Returns the full enriched grid (like GET)."""

    row = await _load_owned_agentic_execution_locked(db, execution_id=execution_id, user=user)
    updated = _set_cell_override(
        row.results,
        document_id=str(payload.document_id),
        column_name=payload.column_name,
        override_value=payload.override_value,
        override_note=payload.override_note,
        overridden_by=str(user.id),
        overridden_at=datetime.now(UTC).isoformat(),
    )
    if updated is None:
        raise HTTPException(status_code=404, detail="cell not found")
    row.results = updated
    # Audit carries IDs/counts only — never the override value or note.
    await audit_action(
        db,
        user_id=user.id,
        action="tabular.cell_overridden",
        resource_type="tabular_execution",
        resource_id=str(execution_id),
        project_id=row.project_id,
        request=request,
        details={"document_id": str(payload.document_id), "column": payload.column_name},
    )
    await db.commit()
    await db.refresh(row)
    response = await _to_response(db, row)
    await _enrich_cell_citations(db, response)
    return response


@router.delete(
    "/tabular/executions/{execution_id}/cells/override",
    response_model=TabularExecutionResponse,
    summary="Clear a lawyer's override on one grid cell (revert to the agent value).",
)
async def clear_tabular_cell_override(
    execution_id: uuid.UUID,
    user: MutatingUser,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
    document_id: Annotated[uuid.UUID, Query()],
    column_name: Annotated[str, Query(min_length=1, max_length=200)],
) -> TabularExecutionResponse:
    """Clear a cell override (revert to the agent value). Idempotent; the
    (document_id, column_name) cell must exist (else 404). Returns the full
    enriched grid (like GET)."""

    row = await _load_owned_agentic_execution_locked(db, execution_id=execution_id, user=user)
    updated = _clear_cell_override(
        row.results,
        document_id=str(document_id),
        column_name=column_name,
    )
    if updated is None:
        raise HTTPException(status_code=404, detail="cell not found")
    row.results = updated
    await audit_action(
        db,
        user_id=user.id,
        action="tabular.cell_override_cleared",
        resource_type="tabular_execution",
        resource_id=str(execution_id),
        project_id=row.project_id,
        request=request,
        details={"document_id": str(document_id), "column": column_name},
    )
    await db.commit()
    await db.refresh(row)
    response = await _to_response(db, row)
    await _enrich_cell_citations(db, response)
    return response


# ---------------------------------------------------------------------------
# Export (M3-C4a — XLSX + CSV)
# ---------------------------------------------------------------------------


_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_CSV_MEDIA_TYPE = "text/csv"


@router.get(
    "/tabular/executions/{execution_id}/export",
    summary="Export a completed tabular execution as XLSX or CSV.",
    responses={
        200: {
            "description": "Streaming export — Content-Type matches format.",
            "content": {
                _XLSX_MEDIA_TYPE: {},
                _CSV_MEDIA_TYPE: {},
            },
        },
        409: {"description": "Execution not in terminal status — cannot export."},
    },
)
async def export_tabular_execution(
    execution_id: uuid.UUID,
    user: ActiveUser,
    request: Request,
    db: Annotated[AsyncSession, Depends(get_db)],
    format: Annotated[
        Literal["xlsx", "csv"],
        Query(
            description="Export format. `xlsx` carries citations as cell comments; `csv` flattens citations into a `citation_links` column."
        ),
    ] = "xlsx",
) -> Response:
    """Stream the tabular grid in XLSX or CSV.

    Both formats include the document column (leftmost) plus one column
    per declared column spec, in spec order. Cells with
    ``confidence='failed'`` render as ``"(failed)"`` rather than an
    empty string so operators can spot the gap without cross-referencing
    the source execution.

    Citation surfacing differs by format:

    * **XLSX** — each grid cell with at least one citation carries an
      openpyxl ``Comment`` listing the citation IDs and confidences.
      Operators can hover any cell in Excel / Numbers / Google Sheets
      to see the source. (Up to 5 citations per comment; the cell
      retains the full count.)
    * **CSV** — a trailing ``citation_links`` column per row carries a
      semicolon-separated list of ``"<column_name>:<citation_id>"``
      pairs across every cell in the row. Empty when no cell in the
      row had citations.

    The execution must be in ``status='completed'``. Pending / running
    / cancelled / failed rows return 409 — partial grids would mislead
    downstream consumers. (DE-XXX candidate: a separate
    ``allow_partial=true`` query param to export partial grids with a
    visible watermark; deferred until an operator surfaces the need.)
    """

    row = await _load_caller_execution(db, execution_id=execution_id, user=user)
    if row.status != "completed":
        raise HTTPException(
            status_code=409,
            detail=(
                f"execution must be in status='completed' to export "
                f"(current status: {row.status!r})"
            ),
        )

    columns = [ColumnSpec.model_validate(c) for c in (row.columns or [])]
    results = TabularResults.model_validate(row.results) if row.results else TabularResults(rows=[])

    await audit_action(
        db,
        user_id=user.id,
        action="tabular.execution_exported",
        resource_type="tabular_execution",
        resource_id=str(execution_id),
        request=request,
        details={"format": format, "row_count": len(results.rows)},
    )
    await db.commit()

    if format == "xlsx":
        content = _build_xlsx(columns=columns, results=results)
        media_type = _XLSX_MEDIA_TYPE
        filename = f"tabular-{execution_id}.xlsx"
    else:
        content = _build_csv(columns=columns, results=results).encode("utf-8")
        media_type = _CSV_MEDIA_TYPE
        filename = f"tabular-{execution_id}.csv"

    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_xlsx(*, columns: list[ColumnSpec], results: TabularResults) -> bytes:
    """Render the grid as an XLSX workbook (openpyxl).

    Cell comments carry citation metadata so an operator can pivot
    from the spreadsheet back to the deployment. The comment text is
    deliberately compact (citation IDs + confidence) — operators
    needing the full source text click through to the deployment URL
    (a deferred enhancement covers stamping a deployment URL on each
    comment, DE-XXX).
    """

    # Lazy import keeps the openpyxl dependency from being imported on
    # the FastAPI app's hot path (export is the only consumer).
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    if ws is None:
        # openpyxl always creates an active sheet on Workbook(); the
        # branch exists for the type checker.
        ws = wb.create_sheet()
    ws.title = "Tabular Review"

    column_names = [c.name for c in columns]

    # Header row.
    ws.cell(row=1, column=1, value="Document")
    for ci, name in enumerate(column_names, start=2):
        ws.cell(row=1, column=ci, value=name)

    # Data rows.
    for ri, grid_row in enumerate(results.rows, start=2):
        ws.cell(row=ri, column=1, value=grid_row.document_name)
        for ci, name in enumerate(column_names, start=2):
            cell_data = grid_row.cells.get(name)
            value = _cell_display_value(cell_data)
            xlsx_cell = ws.cell(row=ri, column=ci, value=value)
            if cell_data and cell_data.citations:
                xlsx_cell.comment = _build_xlsx_comment(cell_data.citations)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_csv(*, columns: list[ColumnSpec], results: TabularResults) -> str:
    """Render the grid as RFC 4180 CSV (stdlib `csv`).

    The trailing ``citation_links`` column flattens every cell's
    citations in the row into a single semicolon-separated string of
    ``<column_name>:<citation_id>`` pairs. Empty when no cell had
    citations.
    """

    out = io.StringIO()
    writer = csv.writer(out)
    column_names = [c.name for c in columns]
    writer.writerow(["Document", *column_names, "citation_links"])

    for grid_row in results.rows:
        row_values: list[str] = [grid_row.document_name]
        row_citations: list[str] = []
        for name in column_names:
            cell_data = grid_row.cells.get(name)
            row_values.append(_cell_display_value(cell_data))
            if cell_data and cell_data.citations:
                row_citations.extend(f"{name}:{c.citation_id}" for c in cell_data.citations)
        row_values.append(";".join(row_citations))
        writer.writerow(row_values)

    return out.getvalue()


def _cell_display_value(cell_data: object) -> str:
    """Project a CellResult into a single string for export rendering.

    Failed cells render as ``"(failed)"`` so the gap is visible in the
    spreadsheet without consulting the source execution. None / missing
    cells render as the empty string.
    """

    if cell_data is None:
        return ""
    confidence = getattr(cell_data, "confidence", None)
    if confidence == "failed":
        return "(failed)"
    value = getattr(cell_data, "value", None)
    return value if value is not None else ""


def _build_xlsx_comment(citations: list[Citation]) -> object:
    """Build an openpyxl Comment summarising the cell's citations.

    Caps at 5 citations to keep the comment readable; the full count
    is preserved in the lead line so operators know whether the
    summary is complete.
    """

    from openpyxl.comments import Comment

    total = len(citations)
    shown = citations[:5]
    lines = [f"{total} citation(s):"]
    for c in shown:
        cid = c.citation_id
        conf = c.confidence
        lines.append(f"- {str(cid)[:8]} ({conf})")
    if total > len(shown):
        lines.append(f"- ... and {total - len(shown)} more")
    return Comment("\n".join(lines), "LQ.AI")


# ---------------------------------------------------------------------------
# Response shaping
# ---------------------------------------------------------------------------


async def _load_caller_execution(
    db: AsyncSession,
    *,
    execution_id: uuid.UUID,
    user: ActiveUser,
) -> TabularExecution:
    """Load one execution + apply the caller-scope visibility rule.

    404 on missing / unauthorized / soft-deleted (single error
    surface; no information leakage).
    """

    row = await db.get(TabularExecution, execution_id)
    if row is None or row.deleted_at is not None:
        raise HTTPException(status_code=404, detail="execution not found")
    if not tenant_admin_visibility(user) and row.user_id != user.id:
        raise HTTPException(status_code=404, detail="execution not found")
    return row


async def _load_document_names(
    db: AsyncSession, document_ids: list[uuid.UUID]
) -> dict[uuid.UUID, str]:
    """Return a ``{document_id: filename}`` map for the given ids.

    Joins ``documents → files.filename``. Soft-deleted files are
    excluded — the caller treats a missing id as a deleted document
    (renders as the empty string in the response, which the UI flags
    as "deleted document")."""

    from app.models.file import File as FileModel

    if not document_ids:
        return {}
    stmt = (
        select(Document.id, FileModel.filename)
        .join(FileModel, Document.file_id == FileModel.id)
        .where(Document.id.in_(document_ids), FileModel.deleted_at.is_(None))
    )
    rows = (await db.execute(stmt)).all()
    return {row.id: row.filename for row in rows}


async def _to_response(db: AsyncSession, row: TabularExecution) -> TabularExecutionResponse:
    """Convert the ORM row to the response wire shape.

    Async because the response includes a ``document_names`` field
    that requires a join on ``files`` — keeping the join inside the
    response builder means every endpoint that returns a single
    execution gets the names without duplicating the query."""

    document_ids = list(row.document_ids)
    name_by_id = await _load_document_names(db, document_ids)
    document_names = [name_by_id.get(did, "") for did in document_ids]
    return TabularExecutionResponse.model_validate(
        {
            "id": row.id,
            "user_id": row.user_id,
            "parent_execution_id": row.parent_execution_id,
            "skill_name": row.skill_name,
            "status": row.status,
            "document_ids": document_ids,
            "document_names": document_names,
            "columns": list(row.columns),
            "results": row.results,
            "cost_estimate_usd": row.cost_estimate_usd,
            "cost_actual_usd": row.cost_actual_usd,
            "error_text": row.error_text,
            "created_at": row.created_at,
            "started_at": row.started_at,
            "completed_at": row.completed_at,
        }
    )


async def _enrich_cell_citations(db: AsyncSession, response: TabularExecutionResponse) -> None:
    """Make tabular cell citations navigable by resolving their source.

    The schema-level ``_synthesize_cell_citations`` validator builds each
    :class:`Citation` purely from the cell's persisted ``cited_chunk_ids``
    (real ``document_chunks.id`` UUIDs) with no DB access, so it can only
    set ``document_id`` (a ``documents.id``) — not enough for the frontend
    to navigate (its doc panel keys off ``files.id``). Here, where a DB
    session is available, we enrich every cell citation in place with
    ``source_file_id`` (``documents.file_id``), ``source_page``
    (``document_chunks.page_start``), and ``source_text``
    (``document_chunks.content``).

    Resolution is batched — two ``IN`` queries total regardless of grid
    size (no N+1): one over every cited chunk id across the whole grid,
    then one over the documents those chunks belong to. Citations whose
    ``chunk_id`` is missing (stale / hard-deleted chunk) are left with the
    navigation fields unset rather than crashing.
    """

    if response.results is None:
        return

    all_citations: list[Citation] = [
        citation
        for tab_row in response.results.rows
        for cell in tab_row.cells.values()
        for citation in cell.citations
    ]
    chunk_ids = {c.chunk_id for c in all_citations if c.chunk_id is not None}
    if not chunk_ids:
        return

    # Query 1: chunk id -> (document_id, page_start, content).
    chunk_stmt = select(
        DocumentChunk.id,
        DocumentChunk.document_id,
        DocumentChunk.page_start,
        DocumentChunk.content,
    ).where(DocumentChunk.id.in_(chunk_ids))
    chunk_rows = (await db.execute(chunk_stmt)).all()
    chunk_by_id = {cr.id: (cr.document_id, cr.page_start, cr.content) for cr in chunk_rows}

    # Query 2: document id -> file_id.
    document_ids = {doc_id for (doc_id, _page, _content) in chunk_by_id.values()}
    file_by_document: dict[uuid.UUID, uuid.UUID] = {}
    if document_ids:
        doc_stmt = select(Document.id, Document.file_id).where(Document.id.in_(document_ids))
        file_by_document = {dr.id: dr.file_id for dr in (await db.execute(doc_stmt)).all()}

    for citation in all_citations:
        if citation.chunk_id is None:
            continue
        resolved = chunk_by_id.get(citation.chunk_id)
        if resolved is None:
            continue  # stale chunk reference — leave navigation fields unset.
        doc_id, page_start, content = resolved
        citation.source_file_id = file_by_document.get(doc_id)
        citation.source_page = page_start
        citation.source_text = content


def _to_summary(row: TabularExecution) -> TabularExecutionSummary:
    """Compact list-endpoint projection."""

    return TabularExecutionSummary.model_validate(
        {
            "id": row.id,
            "user_id": row.user_id,
            "parent_execution_id": row.parent_execution_id,
            "skill_name": row.skill_name,
            "status": row.status,
            "document_count": len(row.document_ids),
            "column_count": len(row.columns),
            "column_names": [
                str(c["name"]) for c in row.columns if isinstance(c, dict) and c.get("name")
            ],
            "fill_mode": row.fill_mode,
            "cost_estimate_usd": row.cost_estimate_usd,
            "cost_actual_usd": row.cost_actual_usd,
            "created_at": row.created_at,
            "completed_at": row.completed_at,
        }
    )


__all__ = ["router"]
