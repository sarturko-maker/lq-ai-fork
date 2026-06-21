"""Article 30 RoPA export — PRIV-4a (fork, ADR-F018/F019).

Pure render layer over the deployment-global ROPA register: turn the read DTOs
(:class:`app.schemas.ropa.ProcessingActivityRead` / ``SystemRead``) into the
Article 30 deliverable in three forms — JSON (machine / queries / re-import),
CSV (spreadsheet, dependency-free), and XLSX (OneTrust's multi-sheet shape).

No DB, no HTTP, no ORM here — the endpoint (``app.api.ropa``) fetches + builds
the DTOs and hands them in, so this module is pure and unit-testable.

**Honest scope.** As of PRIV-6a the domain captures every GDPR Article 30(1)
content field — recipients (PRIV-5a), third-country transfers + safeguards
(PRIV-5b) and the categories-of-data-subjects / categories-of-personal-data
taxonomy (PRIV-6a, Article 30(1)(c)). The coverage note is therefore empty; the
mechanism stays in place (still honest) so a future Article 30(1) gap can
re-populate it. The export RENDERS WHAT EXISTS and never invents fields —
"system proposes, user owns".
"""

from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import TYPE_CHECKING

from app.schemas.ropa import (
    Article30Coverage,
    Article30Export,
    DataCategoryRead,
    DataSubjectCategoryRead,
    ProcessingActivityRead,
    SystemRead,
    VendorRead,
)

if TYPE_CHECKING:  # pragma: no cover - typing only
    from openpyxl.worksheet.worksheet import Worksheet

# Article 30(1) content the domain does not yet model. Surfaced in the export's
# coverage note so the deliverable stays honest about any gap. PRIV-5a filled
# "categories of recipients" (vendors); PRIV-5b filled third-country transfers +
# safeguards; PRIV-6a filled the data-subject / personal-data taxonomy (Article
# 30(1)(c)) — so the register now captures every Article 30(1) content field and
# this tuple is empty. Kept (not deleted) so a future gap can re-populate it.
ART30_FIELDS_NOT_YET_RECORDED: tuple[str, ...] = ()

# Spreadsheet-formula trigger characters (OWASP CSV-injection). The register
# holds model-proposed strings; treat them as untrusted on the way OUT too, so a
# cell like ``=cmd|...`` cannot execute when the file is opened in a spreadsheet.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")

ACTIVITY_HEADER: tuple[str, ...] = (
    "Name",
    "Purpose",
    "Lawful basis (Art 6)",
    "Controller role",
    "Special category",
    "Art 9 condition",
    "Retention",
    "Categories of data subjects (Art 30(1)(c))",
    "Categories of personal data (Art 30(1)(c))",
    "Linked systems",
    "Recipients (Art 30(1)(e))",
    "Transfers (Art 30(1)(e))",
    "Created",
    "Last updated",
)

SYSTEM_HEADER: tuple[str, ...] = (
    "Name",
    "Type",
    "Description",
    "Owner",
    "Hosting location",
    "Retention",
    "Security measures",
    "AI / automated decision-making",
    "Linked processing activities",
    "Created",
    "Last updated",
)

VENDOR_HEADER: tuple[str, ...] = (
    "Name",
    "Role",
    "Country",
    "DPA status",
    "Description",
    "Linked processing activities",
    "Created",
    "Last updated",
)

TRANSFER_HEADER: tuple[str, ...] = (
    "Processing activity",
    "Destination",
    "Restricted (outside UK/EEA)",
    "Transfer mechanism (Chapter V)",
    "Recipient",
    "Safeguard details",
)

DATA_SUBJECT_CATEGORY_HEADER: tuple[str, ...] = (
    "Category of data subjects",
    "Linked processing activities",
)

DATA_CATEGORY_HEADER: tuple[str, ...] = (
    "Category of personal data",
    "Linked processing activities",
)


# System types whose plain capitalisation reads wrong in a lawyer-facing sheet
# (acronyms / hyphenation). Everything else falls back to generic humanisation.
_SYSTEM_TYPE_LABELS = {
    "crm": "CRM",
    "email_marketing": "Email marketing",
    "third_party_processor": "Third-party processor",
}

# Vendor roles whose plain capitalisation reads wrong (hyphenation).
_VENDOR_ROLE_LABELS = {
    "sub_processor": "Sub-processor",
}

# DPA statuses: ``none`` humanises to "None", which in an auditor-facing Article
# 30 sheet — sitting next to genuinely-empty Country/Description cells — reads
# ambiguously (Python null? not recorded?). Spell out the deliberate "no DPA on
# record" state so the deliverable is honest. The rest humanise cleanly.
_DPA_STATUS_LABELS = {
    "none": "No DPA on record",
}

# Transfer mechanisms whose plain capitalisation reads wrong (acronyms / the
# Article reference a lawyer expects). Everything else humanises cleanly.
_TRANSFER_MECHANISM_LABELS = {
    "standard_contractual_clauses": "Standard contractual clauses (SCCs)",
    "uk_idta": "UK IDTA",
    "binding_corporate_rules": "Binding corporate rules (BCRs)",
    "derogation": "Derogation (Art 49)",
}


def _humanize(value: str) -> str:
    """Render an enum value for a human sheet: ``legal_obligation`` → ``Legal obligation``."""
    return value.replace("_", " ").capitalize() if value else value


def _system_type_label(value: str) -> str:
    """Human label for a system type, with acronyms/hyphenation preserved."""
    return _SYSTEM_TYPE_LABELS.get(value, _humanize(value))


def _vendor_role_label(value: str) -> str:
    """Human label for a vendor role, with hyphenation preserved."""
    return _VENDOR_ROLE_LABELS.get(value, _humanize(value))


def _dpa_status_label(value: str) -> str:
    """Human label for a DPA status; spells out ``none`` so the export is unambiguous."""
    return _DPA_STATUS_LABELS.get(value, _humanize(value))


def _transfer_mechanism_label(value: str | None) -> str:
    """Human label for a Chapter V transfer mechanism (acronyms preserved)."""
    if not value:
        return ""
    return _TRANSFER_MECHANISM_LABELS.get(value, _humanize(value))


def _csv_safe(value: str) -> str:
    """Neutralise a spreadsheet-formula trigger by prefixing a single quote (OWASP)."""
    if value and value[0] in _FORMULA_TRIGGERS:
        return "'" + value
    return value


def _date(value: datetime) -> str:
    """Date-only ISO string for human sheets (the time-of-day is noise in a RoPA)."""
    return value.date().isoformat()


def _systems_cell(activity: ProcessingActivityRead) -> str:
    """Join an activity's linked systems into one cell: ``Name (type); Name (type)``."""
    return "; ".join(f"{s.name} ({_system_type_label(s.system_type)})" for s in activity.systems)


def _vendors_cell(activity: ProcessingActivityRead) -> str:
    """Join an activity's recipients into one cell: ``Name (role); Name (role)``."""
    return "; ".join(f"{v.name} ({_vendor_role_label(v.vendor_role)})" for v in activity.vendors)


def _transfers_cell(activity: ProcessingActivityRead) -> str:
    """Join an activity's transfers into one cell.

    Restricted transfers show their safeguard (``Dest — mechanism``); a
    non-restricted (intra-UK/EEA) transfer is marked as such so the absence of a
    mechanism reads as deliberate, not omitted.
    """
    parts: list[str] = []
    for t in activity.transfers:
        if t.restricted:
            parts.append(f"{t.destination} — {_transfer_mechanism_label(t.mechanism)}")
        else:
            parts.append(f"{t.destination} (not restricted)")
    return "; ".join(parts)


def _data_subject_categories_cell(activity: ProcessingActivityRead) -> str:
    """Join an activity's categories of data subjects into one cell."""
    return "; ".join(c.name for c in activity.data_subject_categories)


def _data_categories_cell(activity: ProcessingActivityRead) -> str:
    """Join an activity's categories of personal data into one cell."""
    return "; ".join(c.name for c in activity.data_categories)


def _activities_cell(
    record: SystemRead | VendorRead | DataSubjectCategoryRead | DataCategoryRead,
) -> str:
    """Join a record's linked processing activities into one cell."""
    return "; ".join(a.name for a in record.processing_activities)


def build_export(
    activities: list[ProcessingActivityRead],
    systems: list[SystemRead],
    vendors: list[VendorRead],
    data_subject_categories: list[DataSubjectCategoryRead] | None = None,
    data_categories: list[DataCategoryRead] | None = None,
    *,
    generated_at: datetime,
) -> Article30Export:
    """Assemble the JSON export envelope (the canonical machine form)."""
    return Article30Export(
        generated_at=generated_at,
        coverage=Article30Coverage(fields_not_yet_recorded=list(ART30_FIELDS_NOT_YET_RECORDED)),
        processing_activities=activities,
        systems=systems,
        vendors=vendors,
        data_subject_categories=data_subject_categories or [],
        data_categories=data_categories or [],
    )


def _activity_row(a: ProcessingActivityRead) -> list[str]:
    return [
        a.name,
        a.purpose,
        _humanize(a.lawful_basis),
        _humanize(a.controller_role),
        "Yes" if a.special_category else "No",
        _humanize(a.art9_condition) if a.art9_condition else "",
        a.retention,
        _data_subject_categories_cell(a),
        _data_categories_cell(a),
        _systems_cell(a),
        _vendors_cell(a),
        _transfers_cell(a),
        _date(a.created_at),
        _date(a.updated_at),
    ]


def _category_row(c: DataSubjectCategoryRead | DataCategoryRead) -> list[str]:
    """One taxonomy row: the category name + its linked processing activities."""
    return [c.name, _activities_cell(c)]


def _transfer_rows(export: Article30Export) -> list[list[str]]:
    """Flatten every activity's child transfers into one-row-per-transfer.

    A transfer is a child of its activity (PRIV-5b), so the Transfers sheet is
    built from the already-loaded ``activity.transfers`` — no separate fetch.
    """
    rows: list[list[str]] = []
    for a in export.processing_activities:
        for t in a.transfers:
            rows.append(
                [
                    a.name,
                    t.destination,
                    "Yes" if t.restricted else "No",
                    _transfer_mechanism_label(t.mechanism),
                    t.vendor.name if t.vendor else "",
                    t.details or "",
                ]
            )
    return rows


def _system_row(s: SystemRead) -> list[str]:
    return [
        s.name,
        _system_type_label(s.system_type),
        s.description or "",
        s.owner or "",
        s.hosting_location or "",
        s.retention or "",
        s.security_measures or "",
        "Yes" if s.ai_usage else "No",
        _activities_cell(s),
        _date(s.created_at),
        _date(s.updated_at),
    ]


def _vendor_row(v: VendorRead) -> list[str]:
    return [
        v.name,
        _vendor_role_label(v.vendor_role),
        v.country or "",
        _dpa_status_label(v.dpa_status),
        v.description or "",
        _activities_cell(v),
        _date(v.created_at),
        _date(v.updated_at),
    ]


def to_csv(export: Article30Export) -> str:
    """Render the processing-activities register as CSV (one row per activity).

    A single sheet — the activity is the Article 30 primary record; its systems
    are joined into the ``Linked systems`` cell. Every value is CSV-injection
    guarded.
    """
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(ACTIVITY_HEADER)
    for a in export.processing_activities:
        writer.writerow([_csv_safe(v) for v in _activity_row(a)])
    return buf.getvalue()


def to_xlsx(export: Article30Export) -> bytes:
    """Render a multi-sheet workbook (OneTrust's shape).

    Sheets: Activities + Systems + Vendors + Transfers + Data Subjects + Data
    Categories. ``openpyxl`` is imported lazily so the dependency-free JSON/CSV
    paths never pay for it.
    """
    from openpyxl import Workbook

    wb = Workbook()
    activities_ws = wb.active
    activities_ws.title = "Processing Activities"
    _write_sheet(
        activities_ws,
        ACTIVITY_HEADER,
        [_activity_row(a) for a in export.processing_activities],
    )

    systems_ws = wb.create_sheet("Systems")
    _write_sheet(systems_ws, SYSTEM_HEADER, [_system_row(s) for s in export.systems])

    vendors_ws = wb.create_sheet("Vendors")
    _write_sheet(vendors_ws, VENDOR_HEADER, [_vendor_row(v) for v in export.vendors])

    transfers_ws = wb.create_sheet("Transfers")
    _write_sheet(transfers_ws, TRANSFER_HEADER, _transfer_rows(export))

    data_subjects_ws = wb.create_sheet("Data Subjects")
    _write_sheet(
        data_subjects_ws,
        DATA_SUBJECT_CATEGORY_HEADER,
        [_category_row(c) for c in export.data_subject_categories],
    )

    data_categories_ws = wb.create_sheet("Data Categories")
    _write_sheet(
        data_categories_ws,
        DATA_CATEGORY_HEADER,
        [_category_row(c) for c in export.data_categories],
    )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _write_sheet(ws: Worksheet, header: tuple[str, ...], rows: list[list[str]]) -> None:
    """Write a header + CSV-injection-guarded data rows into a worksheet."""
    ws.append(list(header))
    for row in rows:
        ws.append([_csv_safe(v) for v in row])
