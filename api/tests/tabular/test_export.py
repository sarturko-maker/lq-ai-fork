"""Export helpers for Tabular Review — M3-C4a.

Targets :mod:`app.api.tabular`'s ``_build_xlsx`` and ``_build_csv``
helpers (the body of ``GET /tabular/executions/{id}/export``):

* XLSX round-trips through openpyxl (correct workbook structure +
  cells + comments carrying citation metadata).
* CSV round-trips through stdlib ``csv.reader`` (correct shape; the
  trailing ``citation_links`` column captures every cell's citations).
* Failed cells render as ``"(failed)"`` rather than empty string so
  the gap is visible in the spreadsheet without consulting the source
  execution.
* Empty results (no rows) produce a valid header-only export.
"""

from __future__ import annotations

import csv
import io
import uuid
from decimal import Decimal

from openpyxl import load_workbook

from app.api.tabular import _build_csv, _build_xlsx
from app.schemas.tabular import (
    CellResult,
    Citation,
    ColumnSpec,
    TabularResults,
    TabularRow,
)


def _make_results() -> tuple[list[ColumnSpec], TabularResults]:
    """Two-column x two-row fixture exercising verified + failed cells."""

    columns = [
        ColumnSpec(name="Term", query="What is the term length?"),
        ColumnSpec(name="Survival", query="What survives?"),
    ]
    doc_a = uuid.uuid4()
    doc_b = uuid.uuid4()
    cite_1 = uuid.uuid4()
    cite_2 = uuid.uuid4()
    results = TabularResults(
        rows=[
            TabularRow(
                document_id=doc_a,
                document_name="nda-1-acme-beta.pdf",
                cells={
                    "Term": CellResult(
                        value="3 years from the Effective Date",
                        citations=[
                            Citation(
                                citation_id=cite_1,
                                document_id=doc_a,
                                confidence="high",
                            )
                        ],
                        confidence="high",
                        tier_used=2,
                        cost_usd=Decimal("0.012"),
                    ),
                    "Survival": CellResult(
                        value=None,
                        confidence="failed",
                        error="no citation found",
                    ),
                },
            ),
            TabularRow(
                document_id=doc_b,
                document_name="nda-2-cypress-delta.pdf",
                cells={
                    "Term": CellResult(
                        value="5 years",
                        citations=[
                            Citation(
                                citation_id=cite_2,
                                document_id=doc_b,
                                confidence="medium",
                            )
                        ],
                        confidence="medium",
                        tier_used=2,
                        cost_usd=Decimal("0.015"),
                    ),
                    "Survival": CellResult(
                        value="Confidentiality obligations survive indefinitely.",
                        citations=[],
                        confidence="low",
                        tier_used=2,
                        cost_usd=Decimal("0.008"),
                    ),
                },
            ),
        ]
    )
    return columns, results


def test_xlsx_round_trips_via_openpyxl() -> None:
    """The exported XLSX opens via ``load_workbook`` with the
    documented sheet structure (Document column + one per spec column)
    and per-cell citation comments."""

    columns, results = _make_results()
    xlsx_bytes = _build_xlsx(columns=columns, results=results)

    assert xlsx_bytes[:2] == b"PK", "XLSX is a ZIP container — must start with PK"

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws is not None
    assert ws.title == "Tabular Review"

    # Header row.
    assert ws.cell(row=1, column=1).value == "Document"
    assert ws.cell(row=1, column=2).value == "Term"
    assert ws.cell(row=1, column=3).value == "Survival"

    # First data row — verified Term cell, failed Survival cell.
    assert ws.cell(row=2, column=1).value == "nda-1-acme-beta.pdf"
    assert ws.cell(row=2, column=2).value == "3 years from the Effective Date"
    assert ws.cell(row=2, column=3).value == "(failed)"

    # Citation comment on the verified cell.
    term_cell = ws.cell(row=2, column=2)
    assert term_cell.comment is not None
    assert "1 citation(s)" in term_cell.comment.text
    assert "(high)" in term_cell.comment.text

    # Failed cell has no comment (no citations).
    assert ws.cell(row=2, column=3).comment is None

    # Second data row — both cells populated; only Term has a citation.
    assert ws.cell(row=3, column=1).value == "nda-2-cypress-delta.pdf"
    assert ws.cell(row=3, column=2).value == "5 years"
    assert ws.cell(row=3, column=3).value == (
        "Confidentiality obligations survive indefinitely."
    )
    assert ws.cell(row=3, column=2).comment is not None
    assert ws.cell(row=3, column=3).comment is None


def test_csv_round_trips_via_stdlib_csv() -> None:
    """The exported CSV parses back through ``csv.reader`` to the same
    grid shape, with the trailing ``citation_links`` column flattening
    every cell's citations into a semicolon-separated string."""

    columns, results = _make_results()
    csv_text = _build_csv(columns=columns, results=results)

    rows = list(csv.reader(io.StringIO(csv_text)))
    assert len(rows) == 3  # header + 2 data rows

    # Header.
    assert rows[0] == ["Document", "Term", "Survival", "citation_links"]

    # First data row — failed cell renders as "(failed)"; citation
    # link is "Term:<cite_1>" only (Survival has no citation).
    assert rows[1][0] == "nda-1-acme-beta.pdf"
    assert rows[1][1] == "3 years from the Effective Date"
    assert rows[1][2] == "(failed)"
    assert rows[1][3].startswith("Term:")
    assert ";" not in rows[1][3]  # single citation, no separator

    # Second data row — both Term + Survival populated; Survival has
    # no citations so the citation_links string is just Term's.
    assert rows[2][0] == "nda-2-cypress-delta.pdf"
    assert rows[2][1] == "5 years"
    assert rows[2][2] == "Confidentiality obligations survive indefinitely."
    assert rows[2][3].startswith("Term:")


def test_empty_results_produce_header_only_xlsx() -> None:
    """A completed execution with no rows still exports a valid XLSX
    with the header row (operators see the column spec even when
    no documents resolved)."""

    columns = [ColumnSpec(name="Term", query="What is the term?")]
    empty = TabularResults(rows=[])
    xlsx_bytes = _build_xlsx(columns=columns, results=empty)
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws is not None
    assert ws.cell(row=1, column=1).value == "Document"
    assert ws.cell(row=1, column=2).value == "Term"
    assert ws.cell(row=2, column=1).value is None  # no data row


def test_empty_results_produce_header_only_csv() -> None:
    """Empty results CSV variant."""

    columns = [ColumnSpec(name="Term", query="What is the term?")]
    empty = TabularResults(rows=[])
    csv_text = _build_csv(columns=columns, results=empty)
    rows = list(csv.reader(io.StringIO(csv_text)))
    assert len(rows) == 1
    assert rows[0] == ["Document", "Term", "citation_links"]


def test_xlsx_comment_caps_at_five_citations() -> None:
    """The cell comment caps the citation list at 5 to keep it
    readable, but the total count is preserved in the lead line so
    operators know whether the summary is complete."""

    doc = uuid.uuid4()
    columns = [ColumnSpec(name="Term", query="What is the term?")]
    citations = [
        Citation(citation_id=uuid.uuid4(), document_id=doc, confidence="high")
        for _ in range(7)
    ]
    results = TabularResults(
        rows=[
            TabularRow(
                document_id=doc,
                document_name="seven-citations.pdf",
                cells={
                    "Term": CellResult(
                        value="see citations",
                        citations=citations,
                        confidence="high",
                        tier_used=2,
                        cost_usd=Decimal("0.012"),
                    )
                },
            )
        ]
    )
    xlsx_bytes = _build_xlsx(columns=columns, results=results)
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws is not None
    comment_text = ws.cell(row=2, column=2).comment.text
    assert "7 citation(s)" in comment_text
    assert "and 2 more" in comment_text
