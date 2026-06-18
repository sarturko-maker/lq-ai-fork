"""Article 30 RoPA export — PRIV-4a + PRIV-5a + PRIV-5b (fork, ADR-F018/F019).

Two layers:

* **Pure formatter** (no DB) — the JSON envelope's honest coverage note, the CSV
  header/rows + systems-join + recipients-join + transfers-join cells + the OWASP
  CSV-injection guard, and the four-sheet XLSX workbook (Activities + Systems +
  Vendors + Transfers).
* **Endpoint** (integration) — ``GET /ropa/export`` in each format, the empty
  register, an off-enum ``format`` → 422, and the shared-read auth gate (401).
"""

from __future__ import annotations

import io
import uuid
from collections.abc import AsyncIterator, Callable
from datetime import UTC, datetime

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy import delete
from sqlalchemy.ext.asyncio import AsyncSession

from app import ropa_export
from app.db.session import get_db
from app.main import app
from app.models.ropa import (
    ProcessingActivity,
    System,
    Transfer,
    Vendor,
    processing_activity_systems,
    processing_activity_vendors,
)
from app.models.user import User
from app.schemas.ropa import ProcessingActivityRead, SystemRead, VendorRead
from app.security import create_access_token, hash_password

_NOW = datetime(2026, 6, 18, 9, 30, tzinfo=UTC)


# --- pure formatter (unit) ----------------------------------------------------


def _activity(**over: object) -> ProcessingActivityRead:
    base: dict[str, object] = {
        "id": uuid.uuid4(),
        "name": "Payroll processing",
        "purpose": "Pay employees and meet tax obligations",
        "lawful_basis": "legal_obligation",
        "controller_role": "controller",
        "retention": "7 years",
        "special_category": False,
        "art9_condition": None,
        "created_at": _NOW,
        "updated_at": _NOW,
        "systems": [],
        "vendors": [],
        "transfers": [],
    }
    base.update(over)
    return ProcessingActivityRead.model_validate(base)


def _transfer(**over: object) -> dict[str, object]:
    """A TransferSummary dict as it rides nested under a ProcessingActivityRead."""
    base: dict[str, object] = {
        "id": uuid.uuid4(),
        "destination": "United States",
        "restricted": True,
        "mechanism": "standard_contractual_clauses",
        "details": None,
        "vendor": None,
    }
    base.update(over)
    return base


def _system(**over: object) -> SystemRead:
    base: dict[str, object] = {
        "id": uuid.uuid4(),
        "name": "Production database",
        "system_type": "database",
        "description": None,
        "owner": None,
        "hosting_location": "UK",
        "retention": None,
        "security_measures": None,
        "ai_usage": False,
        "created_at": _NOW,
        "updated_at": _NOW,
        "processing_activities": [],
    }
    base.update(over)
    return SystemRead.model_validate(base)


def _vendor(**over: object) -> VendorRead:
    base: dict[str, object] = {
        "id": uuid.uuid4(),
        "name": "Acme Payroll Ltd",
        "vendor_role": "processor",
        "description": None,
        "country": "UK",
        "dpa_status": "in_place",
        "created_at": _NOW,
        "updated_at": _NOW,
        "processing_activities": [],
    }
    base.update(over)
    return VendorRead.model_validate(base)


@pytest.mark.unit
def test_build_export_carries_honest_coverage_note() -> None:
    export = ropa_export.build_export([_activity()], [_system()], [_vendor()], generated_at=_NOW)
    assert export.generated_at == _NOW
    assert export.register_name == "Article 30 Records of Processing Activities"
    # PRIV-5a filled "categories of recipients"; PRIV-5b filled transfers — neither
    # is in the gap note any more.
    assert "Categories of recipients" not in export.coverage.fields_not_yet_recorded
    assert not any("transfer" in f.lower() for f in export.coverage.fields_not_yet_recorded)
    # The Art 30(1) fields still missing (the data-subject/personal-data taxonomy)
    # are named, not hidden.
    assert any("data subjects" in f.lower() for f in export.coverage.fields_not_yet_recorded)
    assert any("personal data" in f.lower() for f in export.coverage.fields_not_yet_recorded)
    assert len(export.processing_activities) == 1
    assert len(export.systems) == 1
    assert len(export.vendors) == 1


@pytest.mark.unit
def test_to_csv_header_row_and_humanized_values() -> None:
    sysid = uuid.uuid4()
    a = _activity(
        special_category=True,
        art9_condition="employment_social_security",
        systems=[{"id": sysid, "name": "Prod DB", "system_type": "database"}],
    )
    export = ropa_export.build_export([a], [], [], generated_at=_NOW)
    text = ropa_export.to_csv(export)
    lines = text.splitlines()
    assert lines[0].split(",")[0] == "Name"
    assert "Linked systems" in lines[0]
    assert "Recipients" in lines[0]
    # Enum values are humanized for the spreadsheet; the special-category +
    # systems-join cell render as expected.
    assert "Legal obligation" in lines[1]
    assert "Employment social security" in lines[1]
    assert "Prod DB (Database)" in text
    # date-only stamp (no time-of-day noise)
    assert "2026-06-18" in lines[1]
    assert "09:30" not in text


@pytest.mark.unit
def test_recipients_cell_joins_vendors_with_humanized_role() -> None:
    a = _activity(
        vendors=[
            {"id": uuid.uuid4(), "name": "Acme Payroll Ltd", "vendor_role": "processor"},
            {"id": uuid.uuid4(), "name": "SubCo", "vendor_role": "sub_processor"},
        ]
    )
    export = ropa_export.build_export([a], [], [], generated_at=_NOW)
    text = ropa_export.to_csv(export)
    # Recipients joined "Name (role)" with hyphenated sub-processor label.
    assert "Acme Payroll Ltd (Processor)" in text
    assert "SubCo (Sub-processor)" in text


@pytest.mark.unit
def test_system_type_acronyms_read_professionally() -> None:
    # crm → "CRM" (not "Crm"); the join cell + the Systems sheet both use it.
    a = _activity(systems=[{"id": uuid.uuid4(), "name": "Salesforce", "system_type": "crm"}])
    s = _system(name="Salesforce", system_type="crm")
    export = ropa_export.build_export([a], [s], [], generated_at=_NOW)
    assert "Salesforce (CRM)" in ropa_export.to_csv(export)
    wb = load_workbook(io.BytesIO(ropa_export.to_xlsx(export)))
    assert wb["Systems"]["B2"].value == "CRM"


@pytest.mark.unit
def test_to_csv_neutralises_formula_injection() -> None:
    a = _activity(name="=cmd|'/c calc'!A1", purpose="+SUM(1)")
    export = ropa_export.build_export([a], [], [], generated_at=_NOW)
    text = ropa_export.to_csv(export)
    # Both formula-trigger cells are prefixed with a single quote so a
    # spreadsheet won't execute them; the raw "=cmd" never starts a field.
    assert "'=cmd|" in text
    assert "'+SUM(1)" in text


@pytest.mark.unit
def test_to_xlsx_four_sheets_with_headers() -> None:
    a = _activity(
        transfers=[
            _transfer(
                destination="United States",
                vendor={
                    "id": uuid.uuid4(),
                    "name": "Acme Cloud",
                    "vendor_role": "processor",
                },
            ),
        ]
    )
    export = ropa_export.build_export([a], [_system()], [_vendor()], generated_at=_NOW)
    wb = load_workbook(io.BytesIO(ropa_export.to_xlsx(export)))
    assert wb.sheetnames == ["Processing Activities", "Systems", "Vendors", "Transfers"]
    activities = wb["Processing Activities"]
    assert activities["A1"].value == "Name"
    assert activities["A2"].value == "Payroll processing"
    systems = wb["Systems"]
    assert systems["A1"].value == "Name"
    assert systems["A2"].value == "Production database"
    vendors = wb["Vendors"]
    assert vendors["A1"].value == "Name"
    assert vendors["A2"].value == "Acme Payroll Ltd"
    assert vendors["B2"].value == "Processor"
    assert vendors["D2"].value == "In place"
    transfers = wb["Transfers"]
    assert transfers["A1"].value == "Processing activity"
    # One row per transfer, prefixed with its parent activity.
    assert transfers["A2"].value == "Payroll processing"
    assert transfers["B2"].value == "United States"
    assert transfers["C2"].value == "Yes"
    assert transfers["D2"].value == "Standard contractual clauses (SCCs)"
    assert transfers["E2"].value == "Acme Cloud"


@pytest.mark.unit
def test_vendor_none_dpa_status_reads_unambiguously() -> None:
    # ``none`` must not render as the bare "None" (mistakable for a blank cell in
    # an auditor-facing sheet) — it spells out the deliberate no-DPA state.
    export = ropa_export.build_export(
        [], [], [_vendor(name="NoDPA Co", dpa_status="none")], generated_at=_NOW
    )
    wb = load_workbook(io.BytesIO(ropa_export.to_xlsx(export)))
    assert wb["Vendors"]["D2"].value == "No DPA on record"


@pytest.mark.unit
def test_to_xlsx_neutralises_formula_injection_in_vendor_sheet() -> None:
    export = ropa_export.build_export(
        [_activity(name="=danger()")], [], [_vendor(name="@evil")], generated_at=_NOW
    )
    wb = load_workbook(io.BytesIO(ropa_export.to_xlsx(export)))
    assert wb["Processing Activities"]["A2"].value == "'=danger()"
    assert wb["Vendors"]["A2"].value == "'@evil"


@pytest.mark.unit
def test_transfers_cell_shows_restricted_safeguard_and_unrestricted_marker() -> None:
    a = _activity(
        transfers=[
            _transfer(destination="United States", restricted=True, mechanism="uk_idta"),
            _transfer(destination="Germany", restricted=False, mechanism=None),
        ]
    )
    export = ropa_export.build_export([a], [], [], generated_at=_NOW)
    text = ropa_export.to_csv(export)
    assert "Transfers (Art 30(1)(e))" in text.splitlines()[0]
    # A restricted transfer shows its mechanism; a non-restricted one is marked so.
    assert "United States — UK IDTA" in text
    assert "Germany (not restricted)" in text


@pytest.mark.unit
def test_transfer_rows_flatten_across_activities_with_parent_name() -> None:
    a1 = _activity(name="Payroll", transfers=[_transfer(destination="United States")])
    a2 = _activity(name="Marketing", transfers=[_transfer(destination="Singapore")])
    export = ropa_export.build_export([a1, a2], [], [], generated_at=_NOW)
    wb = load_workbook(io.BytesIO(ropa_export.to_xlsx(export)))
    transfers = wb["Transfers"]
    # One row per transfer, each carrying its parent activity name.
    assert transfers["A2"].value == "Payroll"
    assert transfers["B2"].value == "United States"
    assert transfers["A3"].value == "Marketing"
    assert transfers["B3"].value == "Singapore"


@pytest.mark.unit
def test_to_xlsx_neutralises_formula_injection_in_transfer_sheet() -> None:
    a = _activity(transfers=[_transfer(destination="=danger()", restricted=False, mechanism=None)])
    export = ropa_export.build_export([a], [], [], generated_at=_NOW)
    wb = load_workbook(io.BytesIO(ropa_export.to_xlsx(export)))
    assert wb["Transfers"]["B2"].value == "'=danger()"


# --- endpoint (integration) ---------------------------------------------------

pytest_integration = pytest.mark.integration


def _override_get_db(session: AsyncSession) -> Callable[[], AsyncIterator[AsyncSession]]:
    async def _f() -> AsyncIterator[AsyncSession]:
        yield session

    return _f


@pytest_asyncio.fixture
async def client(db_session: AsyncSession) -> AsyncIterator[AsyncClient]:
    app.dependency_overrides[get_db] = _override_get_db(db_session)
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        yield ac
    app.dependency_overrides.pop(get_db, None)


@pytest_asyncio.fixture
async def user(db_session: AsyncSession) -> User:
    u = User(
        email="ropa-export@example.com",
        display_name="ROPA Export",
        hashed_password=hash_password("correct-horse-battery-staple"),
        is_admin=False,
        mfa_enabled=False,
        must_change_password=False,
    )
    db_session.add(u)
    await db_session.flush()
    return u


def _bearer(u: User) -> dict[str, str]:
    token = create_access_token(u.id, u.email, is_admin=u.is_admin)
    return {"Authorization": f"Bearer {token}"}


async def _clean(db_session: AsyncSession) -> None:
    await db_session.execute(delete(processing_activity_systems))
    await db_session.execute(delete(processing_activity_vendors))
    await db_session.execute(delete(Transfer))
    await db_session.execute(delete(ProcessingActivity))
    await db_session.execute(delete(System))
    await db_session.execute(delete(Vendor))
    await db_session.flush()


async def _seed_linked(db_session: AsyncSession) -> None:
    await _clean(db_session)
    pa = ProcessingActivity(
        name="Payroll processing",
        purpose="Pay employees and meet tax obligations",
        lawful_basis="legal_obligation",
        controller_role="controller",
        retention="7 years",
        special_category=False,
        art9_condition=None,
    )
    system = System(name="Production database", system_type="database", hosting_location="UK")
    vendor = Vendor(
        name="Acme Payroll Ltd",
        vendor_role="processor",
        dpa_status="in_place",
        country="UK",
    )
    db_session.add_all([pa, system, vendor])
    await db_session.flush()
    await db_session.execute(
        processing_activity_systems.insert().values(
            processing_activity_id=pa.id, system_id=system.id
        )
    )
    await db_session.execute(
        processing_activity_vendors.insert().values(
            processing_activity_id=pa.id, vendor_id=vendor.id
        )
    )
    db_session.add(
        Transfer(
            processing_activity_id=pa.id,
            vendor_id=vendor.id,
            destination="United States",
            restricted=True,
            mechanism="standard_contractual_clauses",
        )
    )
    await db_session.flush()


@pytest_integration
async def test_export_json(client: AsyncClient, db_session: AsyncSession, user: User) -> None:
    await _seed_linked(db_session)
    resp = await client.get("/api/v1/ropa/export", headers=_bearer(user))
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/json")
    assert "attachment" in resp.headers["content-disposition"]
    assert resp.headers["content-disposition"].endswith('.json"')
    body = resp.json()
    assert body["register_name"] == "Article 30 Records of Processing Activities"
    assert body["coverage"]["fields_not_yet_recorded"]
    assert body["processing_activities"][0]["name"] == "Payroll processing"
    assert body["processing_activities"][0]["systems"][0]["name"] == "Production database"
    assert body["processing_activities"][0]["vendors"][0]["name"] == "Acme Payroll Ltd"
    transfer = body["processing_activities"][0]["transfers"][0]
    assert transfer["destination"] == "United States"
    assert transfer["restricted"] is True
    assert body["systems"][0]["name"] == "Production database"
    assert body["vendors"][0]["name"] == "Acme Payroll Ltd"


@pytest_integration
async def test_export_csv(client: AsyncClient, db_session: AsyncSession, user: User) -> None:
    await _seed_linked(db_session)
    resp = await client.get("/api/v1/ropa/export?format=csv", headers=_bearer(user))
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("text/csv")
    assert resp.headers["content-disposition"].endswith('.csv"')
    lines = resp.text.splitlines()
    assert lines[0].startswith("Name,")
    assert "Payroll processing" in lines[1]
    assert "Production database (Database)" in resp.text
    assert "Acme Payroll Ltd (Processor)" in resp.text
    assert "United States — Standard contractual clauses (SCCs)" in resp.text


@pytest_integration
async def test_export_xlsx(client: AsyncClient, db_session: AsyncSession, user: User) -> None:
    await _seed_linked(db_session)
    resp = await client.get("/api/v1/ropa/export?format=xlsx", headers=_bearer(user))
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert resp.headers["content-disposition"].endswith('.xlsx"')
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["Processing Activities", "Systems", "Vendors", "Transfers"]
    assert wb["Processing Activities"]["A2"].value == "Payroll processing"
    assert wb["Vendors"]["A2"].value == "Acme Payroll Ltd"
    # The seeded restricted transfer appears on the Transfers sheet.
    assert wb["Transfers"]["A2"].value == "Payroll processing"
    assert wb["Transfers"]["B2"].value == "United States"


@pytest_integration
async def test_export_empty_register(
    client: AsyncClient, db_session: AsyncSession, user: User
) -> None:
    await _clean(db_session)
    resp = await client.get("/api/v1/ropa/export", headers=_bearer(user))
    assert resp.status_code == 200
    body = resp.json()
    assert body["processing_activities"] == []
    assert body["systems"] == []
    assert body["vendors"] == []
    # Coverage note is present even when the register is empty (honest scope).
    assert body["coverage"]["fields_not_yet_recorded"]


@pytest_integration
async def test_export_bad_format_is_422(
    client: AsyncClient, db_session: AsyncSession, user: User
) -> None:
    resp = await client.get("/api/v1/ropa/export?format=pdf", headers=_bearer(user))
    assert resp.status_code == 422


@pytest_integration
async def test_export_requires_authentication(client: AsyncClient) -> None:
    resp = await client.get("/api/v1/ropa/export")
    assert resp.status_code == 401
